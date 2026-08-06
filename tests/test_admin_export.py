import unittest
from io import BytesIO

from openpyxl import load_workbook

from src.admin_export import build_users_workbook
from src.database.admin import AdminExportUser


class AdminExportTests(unittest.TestCase):
    def test_builds_readable_excel_file(self) -> None:
        user = AdminExportUser(
            user_id=123,
            username="viewer",
            display_name="Test Viewer",
            is_active=1,
            news_enabled=0,
            started_at="2026-08-01 10:00:00",
            last_started_at="2026-08-02 10:00:00",
            last_activity_at="2026-08-04 12:00:00",
            library_items=5,
            planned_items=1,
            watching_items=1,
            completed_items=2,
            on_hold_items=0,
            rated_items=2,
            tracked_series=1,
        )

        workbook = load_workbook(BytesIO(build_users_workbook((user,))))
        sheet = workbook["Пользователи"]

        self.assertEqual(sheet.cell(2, 1).value, 123)
        self.assertEqual(sheet.cell(2, 4).value, "Да")
        self.assertEqual(sheet.cell(2, 5).value, "Нет")
        self.assertEqual(sheet.cell(2, 14).value, 2)
        self.assertEqual(sheet.cell(2, 15).value, 1)

    def test_saves_formula_like_display_name_as_text(self) -> None:
        for display_name in ("=1+1", "+1+1", "-1+1", "@SUM(A1:A2)"):
            with self.subTest(display_name=display_name):
                user = AdminExportUser(
                    user_id=123,
                    username="viewer",
                    display_name=display_name,
                    is_active=1,
                    news_enabled=0,
                    started_at="2026-08-01 10:00:00",
                    last_started_at="2026-08-02 10:00:00",
                    last_activity_at="2026-08-04 12:00:00",
                    library_items=5,
                    planned_items=1,
                    watching_items=1,
                    completed_items=2,
                    on_hold_items=0,
                    rated_items=2,
                    tracked_series=1,
                )

                workbook = load_workbook(BytesIO(build_users_workbook((user,))))
                cell = workbook["Пользователи"].cell(2, 3)

                self.assertEqual(cell.value, user.display_name)
                self.assertEqual(cell.data_type, "s")
